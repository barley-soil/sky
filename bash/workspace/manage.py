import base64
import json
import os
import urllib
from pathlib import Path
from typing import TypedDict, List
from urllib.parse import quote

from openpyxl import load_workbook


class User(TypedDict):
    """
    Sheet 页 - 0
    """
    # [0] 编号
    code: str
    # [1] 昵称
    nickname: str
    # [2] 端口
    port: int


class Inbound(TypedDict):
    """
    Sheet 页 - 1
    """
    # [0] 编号
    code: str
    # [1] 名称
    name: str
    # [2] 端口
    port: int
    # [3] 密码
    password: str
    # [4] 加密方式
    cipher: str


class Outbound(TypedDict):
    """
    Sheet 页 - 2
    """
    # [0] 编号
    code: str
    # [1] 名称
    name: str
    # [2] 服务器地址
    hostname: str
    # [3] 端口号
    port: int
    # [4] 用户名
    username: str
    # [5] 密码
    password: str


class Link(TypedDict):
    """
    Sheet 页 - 3
    """
    # [0] 编号
    code: str
    # [1] 名称
    name: str
    # [2] 入站代码
    inbound: List[str]
    # [3] 出站代码
    outbound: str
    # [4] 代理地址
    proxy: str
    # [5] 用户名列表
    users: List[str]


class ExternalLink(TypedDict):
    """
    Sheet 页 - 4
    """
    # [0] 编号
    code: str
    # [1] 名称
    name: str
    # [2] 主机地址
    hostname: str
    # [3] 端口
    port: int
    # [4] 密码
    password: str
    # [5] 加密方式
    cipher: str
    # [6] 用户列表
    users: List[str]


class Subscribe(TypedDict):
    # [0] 编号
    code: str
    # [1] 名称
    name: str
    # [2] 主机地址
    hostname: str
    # [3] 密码
    password: str
    # [4] 端口
    port: int
    # [5] 加密算法
    cipher: str


XRAY_JSON = """
{
  "log": {
    "loglevel": "warning",
    "access": "/var/log/xray/access.log",
    "error": "/var/log/xray/error.log"
  },
  "inbounds": [],
  "outbounds": [],
  "routing": {
    "domainStrategy": "AsIs",
    "rules": [
      {
        "type": "field",
        "ip": [
          "0.0.0.0/8",
          "10.0.0.0/8",
          "100.64.0.0/10",
          "127.0.0.0/8",
          "169.254.0.0/16",
          "172.16.0.0/12",
          "192.0.0.0/24",
          "192.0.2.0/24",
          "192.88.99.0/24",
          "192.168.0.0/16",
          "198.18.0.0/15",
          "198.51.100.0/24",
          "203.0.113.0/24",
          "::1/128",
          "fc00::/7",
          "fe80::/10"
        ],
        "outboundTag": "blocked"
      }
    ]
  }
}
"""


def get_subscribe_link(user: User, inbound_list: List[Inbound], link_list: List[Link]) -> List[Subscribe]:
    subscribe_list: List[Subscribe] = []
    for link in [link for link in link_list if user["code"] in link["users"]]:
        for inbound in [inbound for inbound in inbound_list if link["code"] == inbound["code"]]:
            values = link["proxy"].split(":")
            if len(values) == 1:
                hostname = values[0]
                port = inbound["port"]
            else:
                hostname = values[0]
                port = int(values[1])
            subscribe_list.append(Subscribe(
                code=link["code"],
                name=link["name"],
                hostname=hostname,
                password=inbound["password"],
                port=port,
                cipher=inbound["cipher"],
            ))
    return subscribe_list


def nginx_conf(root_path: str, port: int) -> str:
    # 创建目录
    directory = root_path + "/" + str(port)
    Path(directory).mkdir(parents=True, exist_ok=True)

    # Nginx 配置文件
    return """
server {
    listen """ + str(port) + """;

    server_name _;

    root """ + directory + """;

    location / {
        try_files $uri $uri/ =404;
    }

    error_page 500 502 503 504 /50x.html;

    location = /50x.html {
        root html;
    }
}
"""


# noinspection DuplicatedCode
def clash_conf(subscribe_list: List[Subscribe], external_link_list: List[ExternalLink]) -> str:
    clash: List[str] = [
        "port: 7890",
        "socks-port: 7891",
        "redir-port: 7892",
        "allow-lan: false",
        "mode: Rule",
        "log-level: info",
        "proxies:"
    ]
    for it in subscribe_list:
        clash.append(f'''  - name: "{it['name']}"''')
        clash.append('''    type: ss''')
        clash.append(f'''    server: {it['hostname']}''')
        clash.append(f'''    port: {it['port']}''')
        clash.append(f'''    cipher: {it['cipher']}''')
        clash.append(f'''    password: "{it['password']}"''')
        clash.append('''    udp: false''')
    for it in external_link_list:
        clash.append(f'''  - name: "{it['name']}"''')
        clash.append('''    type: ss''')
        clash.append(f'''    server: {it['hostname']}''')
        clash.append(f'''    port: {it['port']}''')
        clash.append(f'''    cipher: {it['cipher']}''')
        clash.append(f'''    password: "{it['password']}"''')
        clash.append('''    udp: false''')
    clash.append("proxy-groups:")
    clash.append('  - name: "Proxy"')
    clash.append('    type: select')
    clash.append('    proxies:')
    clash.append('      - "DIRECT"')
    for it in subscribe_list:
        clash.append(f'''      - "{it['name']}"''')
    for it in external_link_list:
        clash.append(f'''      - "{it['name']}"''')
    clash.append("rules:")
    clash.append("  - IP-CIDR,127.0.0.0/8,DIRECT")
    clash.append("  - IP-CIDR,10.0.0.0/8,DIRECT")
    clash.append("  - IP-CIDR,172.16.0.0/12,DIRECT")
    clash.append("  - IP-CIDR,192.168.0.0/16,DIRECT")
    clash.append("  - IP-CIDR,169.254.0.0/16,DIRECT")
    clash.append("  - IP-CIDR,100.64.0.0/10,DIRECT")
    clash.append("  - GEOIP,CN,DIRECT")
    clash.append("  - MATCH,Proxy")
    return '\n'.join(clash)


# noinspection DuplicatedCode
def ss_conf(subscribe_list: List[Subscribe], external_link_list: List[ExternalLink]) -> str:
    ss: List[str] = []
    for it in subscribe_list:
        server = base64.b64encode(
            f'''{it['cipher']}:{it['password']}@{it['hostname']}:{it['port']}'''.encode('utf-8')
        ).decode('utf-8')
        name_encoded = urllib.parse.quote(it['name'])
        ss.append(f'''ss://{server}#{name_encoded}''')
    for it in external_link_list:
        server = base64.b64encode(
            f'''{it['cipher']}:{it['password']}@{it['hostname']}:{it['port']}'''.encode('utf-8')
        ).decode('utf-8')
        name_encoded = urllib.parse.quote(it['name'])
        ss.append(f'''ss://{server}#{name_encoded}''')
    return '\n'.join(ss)


def subscribe(user_list: List[User], inbound_list: List[Inbound], link_list: List[Link],
              external_link_list: List[ExternalLink]):
    root_path = "/opt"
    nginx_conf_list: List[str] = []
    for user in user_list:
        # 订阅信息
        subscribe_list = get_subscribe_link(user, inbound_list, link_list)

        # Nginx 文件
        nginx_conf_list.append(nginx_conf(root_path, user["port"]))

        # Clash 订阅文件
        with open(f'{root_path}/{user['port']}/clash', "w", encoding="utf-8") as f:
            f.write(clash_conf(subscribe_list, external_link_list))

        # SS 订阅文件
        with open(f'{root_path}/{user['port']}/ss', "w", encoding="utf-8") as f:
            f.write(ss_conf(subscribe_list, external_link_list))

    # Nginx Conf
    with open(f'/etc/nginx/conf.d/default.conf', "w", encoding="utf-8") as f:
        f.write('\n\n'.join(nginx_conf_list))


def xray(inbound_list: List[Inbound], outbound_list: List[Outbound], link_list: List[Link]):
    # Xray JSON
    xray_data = json.loads(XRAY_JSON)

    # 入站
    inbounds = []
    for inbound in inbound_list:
        inbounds.append({
            'port': inbound['port'],
            'protocol': 'shadowsocks',
            'settings': {
                'method': inbound['cipher'],
                'password': inbound['password'],
                'ota': False
            },
            'sniffing': {
                'enabled': True,
                'destOverride': [
                    'http',
                    'tls'
                ]
            },
            'tag': inbound['code']
        })
    xray_data['inbounds'] = inbounds

    # 出站
    outbounds = []
    for outbound in outbound_list:
        outbounds.append({
            'protocol': 'socks',
            'settings': {
                'servers': [{
                    'address': outbound['hostname'],
                    'port': outbound['port'],
                    'users': [{
                        'user': outbound['username'],
                        'pass': outbound['password'],
                    }]
                }]
            },
            'tag': outbound['code']
        })
    xray_data['outbounds'] = outbounds

    # 路由
    for link in link_list:
        xray_data['routing']['rules'].append({
            'type': 'field',
            'inboundTag': link['inbound'],
            'outboundTag': link['outbound']
        })

    # Xray Config
    with open(f'/etc/xray/config.json', "w", encoding="utf-8") as f:
        f.write(json.dumps(xray_data))


def manage(data_file: str):
    # 读取Excel 文件
    wb = load_workbook(data_file, read_only=True, data_only=True)

    # 读取用户数据
    sheet = wb.worksheets[0]
    user_list: List[User] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row is None or row[0] is None:
            break
        user_list.append(User(
            code=row[0].strip(),
            nickname=row[1].strip(),
            port=int(str(row[2]).strip()),
        ))

    # 读取入口数据
    sheet = wb.worksheets[1]
    inbound_list: List[Inbound] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row is None or row[0] is None:
            break
        inbound_list.append(Inbound(
            code=row[0].strip(),
            name=row[1].strip(),
            port=int(str(row[2]).strip()),
            password=row[3].strip(),
            cipher=row[4].strip(),
        ))

    # 读取出口数据
    sheet = wb.worksheets[2]
    outbound_list: List[Outbound] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row is None or row[0] is None:
            break
        outbound_list.append(Outbound(
            code=row[0].strip(),
            name=row[1].strip(),
            hostname=row[2].strip(),
            port=int(str(row[3]).strip()),
            username=row[4].strip(),
            password=row[5].strip()
        ))

    # 读取链路
    sheet = wb.worksheets[3]
    link_list: List[Link] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row is None or row[0] is None:
            break
        link_list.append(Link(
            code=row[0].strip(),
            name=row[1].strip(),
            inbound=row[2].strip().split(","),
            outbound=row[3].strip(),
            proxy=row[4].strip(),
            users=row[5].strip().split(",")
        ))

    # 读取外部链路
    sheet = wb.worksheets[4]
    external_link_list: List[ExternalLink] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row is None or row[0] is None:
            break
        external_link_list.append(ExternalLink(
            code=row[0].strip(),
            name=row[1].strip(),
            hostname=row[2].strip(),
            port=int(str(row[3]).strip()),
            password=row[4].strip(),
            cipher=row[5].strip(),
            users=row[6].strip().split(",")
        ))

    # 构建订阅文件
    subscribe(user_list, inbound_list, link_list, external_link_list)

    # XRay 配置文件
    xray(inbound_list, outbound_list, link_list)

    # 重启服务
    os.system('systemctl restart nginx')
    os.system('systemctl restart xray')


if __name__ == "__main__":
    manage("/home/data.xlsx")
    # manage("/Users/xxscloud/Downloads/data.xlsx")
